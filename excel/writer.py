"""
Módulo de escrita/atualização da planilha Excel do Media Rats - Artgen.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from excel.reader import Solicitacao
from utils.status import STATUS_CORES_EXCEL as STATUS_CORES


class ExcelWriter:
    """Atualiza a planilha Excel após geração de artes.

    Args:
        caminho: Caminho absoluto para o arquivo .xlsx.
    """

    ABA_CONTEUDOS = "CONTEUDOS"
    ABA_AVALIACAO = "AVALIACAO_DETALHADA"

    def __init__(self, caminho: Path) -> None:
        self.caminho = Path(caminho)

    def _encontrar_aba(self, wb, nome: str):
        """Busca aba pelo nome (case-insensitive).

        Args:
            wb: Workbook aberto.
            nome: Nome da aba.

        Returns:
            Worksheet ou None.
        """
        for sn in wb.sheetnames:
            if sn.upper() == nome.upper():
                return wb[sn]
        return None

    def _garantir_aba_avaliacao(self, wb) -> object:
        """Cria aba AVALIACAO_DETALHADA se não existir.

        Args:
            wb: Workbook aberto.

        Returns:
            Worksheet da aba de avaliação.
        """
        ws = self._encontrar_aba(wb, self.ABA_AVALIACAO)
        if ws is not None:
            return ws

        ws = wb.create_sheet(title=self.ABA_AVALIACAO)
        cabecalho = [
            "Protocolo", "Cliente", "Tema", "Numero_Arte",
            "Prompt_Utilizado", "Imagem", "Qualidade",
            "Comentarios", "Usar?", "Data_Revisao",
            "Proximas_Acoes",
        ]
        ws.append(cabecalho)
        header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, _ in enumerate(cabecalho, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        return ws

    def _encontrar_coluna(self, ws, nome: str) -> Optional[int]:
        """Localiza o índice (1-based) de uma coluna pelo nome.

        Args:
            ws: Worksheet.
            nome: Nome da coluna a localizar.

        Returns:
            Índice 1-based ou None se não encontrar.
        """
        for cell in ws[1]:
            if cell.value and str(cell.value).strip().upper() == nome.upper():
                return cell.column
        return None

    def _garantir_coluna(self, ws, nome: str) -> int:
        """Garante que a coluna exista, criando-a se necessário.

        Args:
            ws: Worksheet.
            nome: Nome da coluna.

        Returns:
            Índice 1-based da coluna.
        """
        col_idx = self._encontrar_coluna(ws, nome)
        if col_idx is not None:
            return col_idx
        max_col = ws.max_column or 1
        nova_col = max_col + 1
        header_cell = ws.cell(row=1, column=nova_col)
        header_cell.value = nome
        header_cell.font = Font(bold=True)
        return nova_col

    def atualizar_status(
        self,
        solicitacao: Solicitacao,
        novo_status: str,
        adicionar_data: bool = True,
    ) -> None:
        """Atualiza o status de uma solicitação na aba CONTEUDOS.

        Args:
            solicitacao: Objeto com dados da solicitação (incluindo linha_excel).
            novo_status: Novo valor de status (ex: 'Gerado').
            adicionar_data: Se True, atualiza/cria coluna 'Data_Geracao'.
        """
        wb = openpyxl.load_workbook(self.caminho)
        ws = self._encontrar_aba(wb, self.ABA_CONTEUDOS)
        if ws is None:
            wb.close()
            return

        col_status = self._encontrar_coluna(ws, "STATUS")
        if col_status:
            cell = ws.cell(row=solicitacao.linha_excel, column=col_status)
            cell.value = novo_status
            cor = STATUS_CORES.get(novo_status, "FFFFFF")
            cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")

        if adicionar_data:
            col_data = self._garantir_coluna(ws, "DATA_GERACAO")
            ws.cell(
                row=solicitacao.linha_excel, column=col_data
            ).value = datetime.now().strftime("%d/%m/%Y %H:%M")

        wb.save(self.caminho)
        wb.close()

    def registrar_conclusao(
        self,
        solicitacao: Solicitacao,
        caminhos_imagens: List[str],
    ) -> None:
        """Atualiza status para 'Gerado' e registra avaliação numa única abertura.

        Substitui a chamada sequencial de ``atualizar_status`` + ``registrar_avaliacao``,
        eliminando a abertura dupla do workbook por conclusão de geração.

        Args:
            solicitacao: Objeto com dados da solicitação.
            caminhos_imagens: Lista de caminhos relativos das imagens geradas.
        """
        wb = openpyxl.load_workbook(self.caminho)

        ws_cont = self._encontrar_aba(wb, self.ABA_CONTEUDOS)
        if ws_cont is not None:
            col_status = self._encontrar_coluna(ws_cont, "STATUS")
            if col_status:
                cell = ws_cont.cell(row=solicitacao.linha_excel, column=col_status)
                cell.value = "Gerado"
                cor = STATUS_CORES.get("Gerado", "FFFFFF")
                cell.fill = PatternFill(start_color=cor, end_color=cor, fill_type="solid")
            col_data = self._garantir_coluna(ws_cont, "DATA_GERACAO")
            ws_cont.cell(
                row=solicitacao.linha_excel, column=col_data
            ).value = datetime.now().strftime("%d/%m/%Y %H:%M")

        ws_aval = self._garantir_aba_avaliacao(wb)
        hoje = date.today().strftime("%d/%m/%Y")
        for i, caminho_img in enumerate(caminhos_imagens, start=1):
            prompt = solicitacao.prompts[i - 1] if i - 1 < len(solicitacao.prompts) else ""
            ws_aval.append([
                solicitacao.protocolo,
                solicitacao.cliente,
                solicitacao.tema,
                i,
                prompt,
                caminho_img,
                "",
                "",
                "",
                hoje,
                "",
            ])

        wb.save(self.caminho)
        wb.close()

    def remover_solicitacao(self, solicitacao: Solicitacao) -> bool:
        """Remove permanentemente uma solicitação da aba CONTEUDOS.

        Localiza a linha pelo protocolo (mais seguro que linha_excel, pois
        linhas podem ter sido deslocadas por remoções anteriores) e a
        deleta fisicamente da planilha.

        Args:
            solicitacao: Objeto com dados da solicitação a remover.

        Returns:
            True se a linha foi encontrada e removida.

        Raises:
            IOError: Se não for possível salvar a planilha.
        """
        wb = openpyxl.load_workbook(self.caminho)
        ws = self._encontrar_aba(wb, self.ABA_CONTEUDOS)
        if ws is None:
            wb.close()
            return False

        col_protocolo = self._encontrar_coluna(ws, "PROTOCOLO")
        if col_protocolo is None:
            wb.close()
            return False

        linha_alvo = None
        for row_idx in range(2, ws.max_row + 1):
            celula = ws.cell(row=row_idx, column=col_protocolo)
            if celula.value and str(celula.value).strip() == solicitacao.protocolo:
                linha_alvo = row_idx
                break

        if linha_alvo is None:
            wb.close()
            return False

        ws.delete_rows(linha_alvo, 1)
        wb.save(self.caminho)
        wb.close()
        return True

    def registrar_avaliacao(
        self,
        solicitacao: Solicitacao,
        caminhos_imagens: List[str],
    ) -> None:
        """Cria/atualiza linhas de avaliação para os arquivos gerados.

        Args:
            solicitacao: Objeto com dados da solicitação.
            caminhos_imagens: Lista de caminhos relativos das imagens geradas.
        """
        wb = openpyxl.load_workbook(self.caminho)
        ws_aval = self._garantir_aba_avaliacao(wb)

        hoje = date.today().strftime("%d/%m/%Y")

        for i, caminho_img in enumerate(caminhos_imagens, start=1):
            prompt = solicitacao.prompts[i - 1] if i - 1 < len(solicitacao.prompts) else ""
            linha = [
                solicitacao.protocolo,
                solicitacao.cliente,
                solicitacao.tema,
                i,
                prompt,
                caminho_img,
                "",
                "",
                "",
                hoje,
                "",
            ]
            ws_aval.append(linha)

        wb.save(self.caminho)
        wb.close()

    _COLUNAS_CLIENTES = [
        "CODIGO_CLIENTE", "NOME", "NICHO", "DESCRICAO", "PUBLICO_ALVO",
        "FORMALIDADE", "ESTILO_VISUAL", "ESTILO_FOTO",
        "COR_PRIMARIA", "COR_SECUNDARIA", "COR_FUNDO",
    ]

    def _garantir_cabecalho_clientes(self, ws) -> dict:
        """Garante que todas as colunas de CLIENTES existam; retorna mapa nome->col 1-based.

        Args:
            ws: Worksheet CLIENTES.

        Returns:
            Dicionário {NOME_COLUNA: indice_1based}.
        """
        cab = {}
        if ws.max_row < 1:
            ws.append(self._COLUNAS_CLIENTES)

        for cell in ws[1]:
            if cell.value:
                cab[str(cell.value).strip().upper()] = cell.column

        for nome_col in self._COLUNAS_CLIENTES:
            if nome_col not in cab:
                nova_col = (ws.max_column or 0) + 1
                cell = ws.cell(row=1, column=nova_col)
                cell.value = nome_col
                cell.font = Font(bold=True)
                cab[nome_col] = nova_col
        return cab

    def listar_clientes(self) -> list:
        """Retorna todos os clientes da aba CLIENTES como lista de dicts, em ordem alfabética.

        Returns:
            Lista de dicts com os campos do cliente.
        """
        wb = openpyxl.load_workbook(self.caminho)
        ws = self._encontrar_aba(wb, "CLIENTES")
        clientes = []
        if ws:
            cab = {}
            primeira = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if primeira:
                for i, v in enumerate(primeira):
                    if v:
                        cab[str(v).strip().upper()] = i

            def _get(row, *chaves):
                for ch in chaves:
                    idx = cab.get(ch.upper())
                    if idx is not None and idx < len(row) and row[idx]:
                        return str(row[idx]).strip()
                return ""

            for row in ws.iter_rows(min_row=2, values_only=True):
                if row and row[0]:
                    codigo = str(row[0]).strip().upper()
                    clientes.append({
                        "codigo": codigo,
                        "nome": _get(row, "NOME", "NOME_CLIENTE") or codigo,
                        "nicho": _get(row, "NICHO"),
                        "descricao": _get(row, "DESCRICAO"),
                        "publico_alvo": _get(row, "PUBLICO_ALVO"),
                        "formalidade": _get(row, "FORMALIDADE"),
                        "estilo_visual": _get(row, "ESTILO_VISUAL"),
                        "estilo_foto": _get(row, "ESTILO_FOTO"),
                        "cor_primaria": _get(row, "COR_PRIMARIA"),
                        "cor_secundaria": _get(row, "COR_SECUNDARIA"),
                        "cor_fundo": _get(row, "COR_FUNDO"),
                    })
        wb.close()
        clientes.sort(key=lambda c: c["nome"].lower())
        return clientes

    def adicionar_cliente(
        self,
        codigo: str,
        nome: str,
        nicho: str = "",
        descricao: str = "",
        publico_alvo: str = "",
        formalidade: str = "",
        estilo_visual: str = "",
        estilo_foto: str = "",
        cor_primaria: str = "",
        cor_secundaria: str = "",
        cor_fundo: str = "",
    ) -> None:
        """Adiciona um novo cliente na aba CLIENTES.

        Args:
            codigo: Código único do cliente (2-6 letras).
            nome: Nome do cliente.
            nicho: Nicho de mercado.
            descricao: Descrição da empresa.
            publico_alvo: Público-alvo.
            formalidade: Nível de formalidade.
            estilo_visual: Estilo visual.
            estilo_foto: Estilo fotográfico.
            cor_primaria: Cor primária em HEX.
            cor_secundaria: Cor secundária em HEX.
            cor_fundo: Cor de fundo em HEX.

        Raises:
            ValueError: Se o código já existir.
        """
        wb = openpyxl.load_workbook(self.caminho)
        ws = self._encontrar_aba(wb, "CLIENTES")
        if ws is None:
            ws = wb.create_sheet("CLIENTES")

        cab = self._garantir_cabecalho_clientes(ws)
        codigo = codigo.strip().upper()

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0] and str(row[0]).strip().upper() == codigo:
                wb.close()
                raise ValueError(f"Cliente com código '{codigo}' já existe.")

        nova_linha = ws.max_row + 1
        dados = {
            "CODIGO_CLIENTE": codigo,
            "NOME": nome.strip(),
            "NICHO": nicho.strip(),
            "DESCRICAO": descricao.strip(),
            "PUBLICO_ALVO": publico_alvo.strip(),
            "FORMALIDADE": formalidade.strip(),
            "ESTILO_VISUAL": estilo_visual.strip(),
            "ESTILO_FOTO": estilo_foto.strip(),
            "COR_PRIMARIA": cor_primaria.strip(),
            "COR_SECUNDARIA": cor_secundaria.strip(),
            "COR_FUNDO": cor_fundo.strip(),
        }
        for col_nome, col_idx in cab.items():
            if col_nome in dados:
                ws.cell(row=nova_linha, column=col_idx).value = dados[col_nome]

        wb.save(self.caminho)
        wb.close()

    def atualizar_cliente(
        self,
        codigo_original: str,
        novo_codigo: str,
        novo_nome: str,
        nicho: str = "",
        descricao: str = "",
        publico_alvo: str = "",
        formalidade: str = "",
        estilo_visual: str = "",
        estilo_foto: str = "",
        cor_primaria: str = "",
        cor_secundaria: str = "",
        cor_fundo: str = "",
    ) -> None:
        """Atualiza os dados de um cliente existente.

        Args:
            codigo_original: Código atual do cliente.
            novo_codigo: Novo código (pode ser igual ao original).
            novo_nome: Novo nome do cliente.
            Demais args: campos estendidos do cadastro.

        Raises:
            ValueError: Se o cliente não for encontrado.
        """
        wb = openpyxl.load_workbook(self.caminho)
        ws = self._encontrar_aba(wb, "CLIENTES")
        if ws is None:
            wb.close()
            raise ValueError("Aba CLIENTES não encontrada.")

        cab = self._garantir_cabecalho_clientes(ws)
        codigo_original = codigo_original.strip().upper()
        novo_codigo = novo_codigo.strip().upper()

        dados = {
            "CODIGO_CLIENTE": novo_codigo,
            "NOME": novo_nome.strip(),
            "NICHO": nicho.strip(),
            "DESCRICAO": descricao.strip(),
            "PUBLICO_ALVO": publico_alvo.strip(),
            "FORMALIDADE": formalidade.strip(),
            "ESTILO_VISUAL": estilo_visual.strip(),
            "ESTILO_FOTO": estilo_foto.strip(),
            "COR_PRIMARIA": cor_primaria.strip(),
            "COR_SECUNDARIA": cor_secundaria.strip(),
            "COR_FUNDO": cor_fundo.strip(),
        }

        for row in ws.iter_rows(min_row=2):
            if row[0].value and str(row[0].value).strip().upper() == codigo_original:
                for col_nome, col_idx in cab.items():
                    if col_nome in dados:
                        ws.cell(row=row[0].row, column=col_idx).value = dados[col_nome]
                wb.save(self.caminho)
                wb.close()
                return

        wb.close()
        raise ValueError(f"Cliente '{codigo_original}' não encontrado.")

    def remover_cliente(self, codigo: str) -> None:
        """Remove um cliente da aba CLIENTES.

        Args:
            codigo: Código do cliente a remover.

        Raises:
            ValueError: Se o cliente não for encontrado.
        """
        wb = openpyxl.load_workbook(self.caminho)
        ws = self._encontrar_aba(wb, "CLIENTES")
        if ws is None:
            wb.close()
            raise ValueError("Aba CLIENTES não encontrada.")

        codigo = codigo.strip().upper()
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=1)
            if cell.value and str(cell.value).strip().upper() == codigo:
                ws.delete_rows(row_idx)
                wb.save(self.caminho)
                wb.close()
                return

        wb.close()
        raise ValueError(f"Cliente '{codigo}' não encontrado.")

    def proximo_numero_solicitacao(self, codigo_cliente: str) -> int:
        """Retorna o próximo número sequencial de solicitação para um cliente.

        Args:
            codigo_cliente: Código do cliente.

        Returns:
            Próximo número inteiro disponível (começa em 1).
        """
        wb = openpyxl.load_workbook(self.caminho)
        ws = self._encontrar_aba(wb, self.ABA_CONTEUDOS)
        if ws is None:
            wb.close()
            return 1

        cab = {}
        primeira = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if primeira:
            for i, v in enumerate(primeira):
                if v:
                    cab[str(v).strip().upper()] = i

        codigo_cliente = codigo_cliente.strip().upper()
        col_cod = cab.get("CODIGO_CLIENTE", cab.get("CODIGO", None))
        col_num = cab.get("NUMERO_SOLICITACAO", cab.get("NUMERO", None))

        maior = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            cod_val = str(row[col_cod]).strip().upper() if col_cod is not None and col_cod < len(row) and row[col_cod] else ""
            if cod_val != codigo_cliente:
                continue
            if col_num is not None and col_num < len(row) and row[col_num]:
                try:
                    n = int(float(str(row[col_num])))
                    if n > maior:
                        maior = n
                except (ValueError, TypeError):
                    pass
        wb.close()
        return maior + 1

    def adicionar_solicitacao(
        self,
        codigo_cliente: str,
        nome_cliente: str,
        tema: str,
        prompts: list,
        status: str = "Planejado",
        data_planejada: str = "",
    ) -> "Solicitacao":
        """Insere uma nova solicitação na aba CONTEUDOS e retorna o objeto criado.

        Args:
            codigo_cliente: Código do cliente.
            nome_cliente: Nome do cliente.
            tema: Tema da solicitação.
            prompts: Lista de prompts (até 10).
            status: Status inicial.
            data_planejada: Data opcional em 'DD/MM/YYYY'.

        Returns:
            Objeto Solicitacao com linha_excel preenchida.
        """
        from excel.reader import Solicitacao

        wb = openpyxl.load_workbook(self.caminho)
        ws = self._encontrar_aba(wb, self.ABA_CONTEUDOS)
        if ws is None:
            ws = wb.create_sheet(self.ABA_CONTEUDOS)
            cabecalho_cont = (
                ["PROTOCOLO", "CODIGO_CLIENTE", "CLIENTE", "NUMERO_SOLICITACAO",
                 "TEMA", "STATUS", "DATA_PLANEJADA"]
                + [f"PROMPT {i}" for i in range(1, 11)]
            )
            ws.append(cabecalho_cont)
            header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
            for col_idx in range(1, len(cabecalho_cont) + 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
        wb.close()

        numero = self.proximo_numero_solicitacao(codigo_cliente)
        protocolo = f"{codigo_cliente.strip().upper()}#{numero}"
        data_criacao = data_planejada or datetime.now().strftime("%d/%m/%Y")

        wb = openpyxl.load_workbook(self.caminho)
        ws = self._encontrar_aba(wb, self.ABA_CONTEUDOS)

        cab = {}
        primeira = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if primeira:
            for i, v in enumerate(primeira):
                if v:
                    cab[str(v).strip().upper()] = i + 1

        nova_linha = ws.max_row + 1
        campos = {
            "PROTOCOLO": protocolo,
            "CODIGO_CLIENTE": codigo_cliente.strip().upper(),
            "CLIENTE": nome_cliente.strip(),
            "NUMERO_SOLICITACAO": numero,
            "TEMA": tema.strip(),
            "STATUS": status,
            "DATA_PLANEJADA": data_criacao,
        }
        for i, p in enumerate(prompts[:10], start=1):
            campos[f"PROMPT {i}"] = p

        for col_nome, col_idx in cab.items():
            if col_nome in campos:
                cell = ws.cell(row=nova_linha, column=col_idx)
                cell.value = campos[col_nome]

        cor = STATUS_CORES.get(status, "FFFFFF")
        col_status = cab.get("STATUS")
        if col_status:
            ws.cell(row=nova_linha, column=col_status).fill = PatternFill(
                start_color=cor, end_color=cor, fill_type="solid"
            )

        wb.save(self.caminho)
        wb.close()

        prompts_completos = list(prompts[:10]) + [""] * (10 - len(prompts[:10]))
        return Solicitacao(
            linha_excel=nova_linha,
            codigo_cliente=codigo_cliente.strip().upper(),
            cliente=nome_cliente.strip(),
            numero_solicitacao=numero,
            protocolo=protocolo,
            tema=tema.strip(),
            prompts=prompts_completos,
            status=status,
        )

    def criar_estrutura_planilha(self) -> None:
        """Cria as abas e cabeçalhos mínimos se a planilha estiver vazia.
        Não sobrescreve dados existentes.
        """
        wb = openpyxl.load_workbook(self.caminho)
        abas_existentes = [s.upper() for s in wb.sheetnames]

        if "CLIENTES" not in abas_existentes:
            ws_cli = wb.create_sheet("CLIENTES")
            ws_cli.append(self._COLUNAS_CLIENTES)
            header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
            for col_idx in range(1, len(self._COLUNAS_CLIENTES) + 1):
                cell = ws_cli.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")
        else:
            ws_cli = self._encontrar_aba(wb, "CLIENTES")
            if ws_cli is not None:
                self._garantir_cabecalho_clientes(ws_cli)

        if self.ABA_CONTEUDOS.upper() not in abas_existentes:
            ws_cont = wb.create_sheet(self.ABA_CONTEUDOS)
            cabecalho_cont = (
                ["PROTOCOLO", "CODIGO_CLIENTE", "CLIENTE", "NUMERO_SOLICITACAO", "TEMA", "STATUS",
                 "DATA_PLANEJADA"]
                + [f"PROMPT {i}" for i in range(1, 11)]
            )
            ws_cont.append(cabecalho_cont)
            header_fill = PatternFill(start_color="1A237E", end_color="1A237E", fill_type="solid")
            for col_idx in range(1, len(cabecalho_cont) + 1):
                cell = ws_cont.cell(row=1, column=col_idx)
                cell.fill = header_fill
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal="center")

        self._garantir_aba_avaliacao(wb)
        wb.save(self.caminho)
        wb.close()
