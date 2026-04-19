"""
Módulo de leitura da planilha Excel do Media Rats - Artgen.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import List, Optional

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@dataclass
class Cliente:
    """Representa um cliente cadastrado na aba CLIENTES.

    Attributes:
        codigo: Código único do cliente (2-6 letras).
        nome: Nome do cliente.
        nicho: Nicho de mercado.
        descricao: Descrição do cliente/empresa.
        publico_alvo: Público-alvo principal.
        formalidade: Nível de formalidade (ex: Formal, Casual, Técnico).
        estilo_visual: Estilo visual predominante.
        estilo_foto: Estilo fotográfico preferido.
        cor_primaria: Cor primária em HEX (ex: #1565C0).
        cor_secundaria: Cor secundária em HEX.
        cor_fundo: Cor de fundo preferida em HEX.
    """

    codigo: str
    nome: str
    nicho: str = ""
    descricao: str = ""
    publico_alvo: str = ""
    formalidade: str = ""
    estilo_visual: str = ""
    estilo_foto: str = ""
    cor_primaria: str = ""
    cor_secundaria: str = ""
    cor_fundo: str = ""


@dataclass
class Solicitacao:
    """Representa uma solicitação de geração de artes.

    Attributes:
        linha_excel: Número da linha na planilha (para atualização).
        codigo_cliente: Código do cliente (ex: DUDE).
        cliente: Nome do cliente.
        numero_solicitacao: Número sequencial da solicitação.
        protocolo: Protocolo gerado (ex: DUDE#1).
        tema: Tema da solicitação.
        prompts: Lista com até 10 prompts.
        status: Status atual da solicitação.
        data_planejada: Data planejada (opcional).
    """

    linha_excel: int
    codigo_cliente: str
    cliente: str
    numero_solicitacao: int
    protocolo: str
    tema: str
    prompts: List[str] = field(default_factory=list)
    status: str = "Planejado"
    data_planejada: Optional[date] = None

    def prompts_validos(self) -> List[str]:
        """Retorna apenas os prompts não-vazios.

        Returns:
            Lista de prompts preenchidos.
        """
        return [p for p in self.prompts if p and str(p).strip()]


class ExcelReader:
    """Lê e valida a planilha de artes do Media Rats.

    Args:
        caminho: Caminho absoluto para o arquivo .xlsx.
    """

    ABA_CLIENTES = "CLIENTES"
    ABA_CONTEUDOS = "CONTEUDOS"
    ABA_AVALIACAO = "AVALIACAO_DETALHADA"

    STATUS_GERAR = {"planejado", "pendente"}

    def __init__(self, caminho: Path) -> None:
        self.caminho = Path(caminho)
        self._wb: Optional[Workbook] = None

    def _abrir(self) -> Workbook:
        """Abre o arquivo Excel.

        Returns:
            Workbook aberto.

        Raises:
            FileNotFoundError: Se o arquivo não existir.
            ValueError: Se as abas obrigatórias não forem encontradas.
        """
        if not self.caminho.exists():
            raise FileNotFoundError(f"Planilha não encontrada: {self.caminho}")
        wb = openpyxl.load_workbook(self.caminho, data_only=True, read_only=True)
        abas_existentes = [s.upper() for s in wb.sheetnames]
        for aba in [self.ABA_CLIENTES, self.ABA_CONTEUDOS]:
            if aba not in abas_existentes:
                raise ValueError(
                    f"Aba obrigatória '{aba}' não encontrada na planilha. "
                    f"Abas disponíveis: {wb.sheetnames}"
                )
        return wb

    def _encontrar_aba(self, wb: Workbook, nome: str) -> Optional[Worksheet]:
        """Busca aba pelo nome (case-insensitive).

        Args:
            wb: Workbook aberto.
            nome: Nome da aba.

        Returns:
            Worksheet ou None se não encontrar.
        """
        for sheet_name in wb.sheetnames:
            if sheet_name.upper() == nome.upper():
                return wb[sheet_name]
        return None

    def ler_clientes(self) -> List[Cliente]:
        """Lê todos os clientes da aba CLIENTES.

        Usa mapa de cabeçalho para independência de ordem de colunas.
        Retorna clientes em ordem alfabética por nome.

        Returns:
            Lista de objetos Cliente.
        """
        wb = self._abrir()
        ws = self._encontrar_aba(wb, self.ABA_CLIENTES)
        clientes: List[Cliente] = []
        if ws is None:
            wb.close()
            return clientes

        cab = {}
        primeira = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if primeira:
            for i, v in enumerate(primeira):
                if v:
                    cab[str(v).strip().upper()] = i

        def _get(row, *chaves):
            for ch in chaves:
                idx = cab.get(ch.upper())
                if idx is not None and idx < len(row) and row[idx]:
                    return str(row[idx]).strip()
            return ""

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            codigo = str(row[0]).strip().upper()
            if len(codigo) < 2:
                continue
            clientes.append(Cliente(
                codigo=codigo,
                nome=_get(row, "NOME", "NOME_CLIENTE") or codigo,
                nicho=_get(row, "NICHO"),
                descricao=_get(row, "DESCRICAO", "DESCRIÇÃO"),
                publico_alvo=_get(row, "PUBLICO_ALVO", "PÚBLICO_ALVO"),
                formalidade=_get(row, "FORMALIDADE", "NIVEL_FORMALIDADE"),
                estilo_visual=_get(row, "ESTILO_VISUAL"),
                estilo_foto=_get(row, "ESTILO_FOTO", "ESTILO_FOTOGRAFICO"),
                cor_primaria=_get(row, "COR_PRIMARIA", "COR_PRIMÁRIA"),
                cor_secundaria=_get(row, "COR_SECUNDARIA", "COR_SECUNDÁRIA"),
                cor_fundo=_get(row, "COR_FUNDO"),
            ))

        wb.close()
        clientes.sort(key=lambda c: c.nome.lower())
        return clientes

    def ler_solicitacoes(self, apenas_pendentes: bool = True) -> List[Solicitacao]:
        """Lê solicitações da aba CONTEUDOS.

        Args:
            apenas_pendentes: Se True, retorna apenas status Planejado/Pendente.

        Returns:
            Lista de objetos Solicitacao.
        """
        wb = self._abrir()
        ws = self._encontrar_aba(wb, self.ABA_CONTEUDOS)
        solicitacoes: List[Solicitacao] = []
        if ws is None:
            wb.close()
            return solicitacoes

        cabecalho = self._ler_cabecalho(ws)

        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not row[0]:
                continue

            try:
                sol = self._parsear_linha(row, idx, cabecalho)
            except Exception as exc:
                raise ValueError(
                    f"Erro ao ler linha {idx} da aba CONTEUDOS: {exc}"
                ) from exc

            if apenas_pendentes:
                if sol.status.lower() not in self.STATUS_GERAR:
                    continue

            solicitacoes.append(sol)

        wb.close()
        return solicitacoes

    def _ler_cabecalho(self, ws: Worksheet) -> dict:
        """Mapeia nomes de colunas para seus índices (0-based).

        Args:
            ws: Worksheet da planilha.

        Returns:
            Dicionário {nome_coluna_upper: indice}.
        """
        cabecalho = {}
        primeira_linha = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if primeira_linha:
            for i, val in enumerate(primeira_linha):
                if val:
                    cabecalho[str(val).strip().upper()] = i
        return cabecalho

    def _parsear_linha(self, row: tuple, linha_idx: int, cabecalho: dict) -> Solicitacao:
        """Converte uma linha da planilha em objeto Solicitacao.

        Args:
            row: Tupla com os valores da linha.
            linha_idx: Índice da linha no Excel.
            cabecalho: Mapeamento de colunas.

        Returns:
            Objeto Solicitacao preenchido.
        """

        def get(nome: str, padrao=""):
            idx = cabecalho.get(nome.upper())
            if idx is not None and idx < len(row):
                v = row[idx]
                return str(v).strip() if v is not None else padrao
            return padrao

        def get_int(nome: str, padrao: int = 0) -> int:
            val = get(nome, str(padrao))
            try:
                return int(float(val))
            except (ValueError, TypeError):
                return padrao

        codigo_cliente = get("CODIGO_CLIENTE") or get("CODIGO") or get("CLIENTE_CODIGO", "XX")
        codigo_cliente = codigo_cliente.upper()[:4]

        cliente = get("CLIENTE") or get("NOME_CLIENTE", codigo_cliente)
        numero = get_int("NUMERO_SOLICITACAO") or get_int("NUMERO", linha_idx - 1)
        if numero == 0:
            numero = linha_idx - 1

        protocolo_raw = get("PROTOCOLO", "")
        if protocolo_raw:
            protocolo = protocolo_raw
        else:
            protocolo = f"{codigo_cliente}#{numero}"

        tema = get("TEMA", "Sem Tema")
        status = get("STATUS", "Planejado")

        data_str = get("DATA_PLANEJADA") or get("DATA PLANEJADA", "")
        data_planejada = None
        if data_str:
            for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
                try:
                    data_planejada = datetime.strptime(data_str, fmt).date()
                    break
                except ValueError:
                    continue

        prompts: List[str] = []
        for i in range(1, 11):
            chaves = [f"PROMPT {i}", f"PROMPT{i}", f"PROMPT_{i}"]
            for chave in chaves:
                val = get(chave, "")
                if val:
                    prompts.append(val)
                    break
            else:
                prompts.append("")

        return Solicitacao(
            linha_excel=linha_idx,
            codigo_cliente=codigo_cliente,
            cliente=cliente,
            numero_solicitacao=numero,
            protocolo=protocolo,
            tema=tema,
            prompts=prompts,
            status=status,
            data_planejada=data_planejada,
        )

    def verificar_planilha(self) -> List[str]:
        """Executa validações na planilha e retorna lista de erros/avisos.

        Returns:
            Lista de strings com problemas encontrados.
        """
        erros: List[str] = []
        if not self.caminho.exists():
            erros.append(f"Arquivo não encontrado: {self.caminho}")
            return erros
        try:
            wb = self._abrir()
            wb.close()
        except ValueError as e:
            erros.append(str(e))
        return erros
