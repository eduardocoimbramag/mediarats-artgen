"""
Script utilitário para criar/recriar a estrutura da planilha Excel do Artgen.
Execute: python setup_planilha.py
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
PLANILHA_PATH = BASE_DIR / "planilha" / "planilha-artgenmediarats.xlsx"


def estilo_cabecalho(ws, linha: int, cor_fundo: str = "1A237E") -> None:
    """Aplica estilo de cabeçalho a uma linha inteira.

    Args:
        ws: Worksheet alvo.
        linha: Número da linha (1-based).
        cor_fundo: Cor hexadecimal do fundo.
    """
    for cell in ws[linha]:
        if cell.value is None:
            continue
        cell.font = Font(bold=True, color="FFFFFF", name="Segoe UI", size=10)
        cell.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def ajustar_colunas(ws, larguras: dict) -> None:
    """Define largura das colunas.

    Args:
        ws: Worksheet alvo.
        larguras: Dict {letra_coluna: largura}.
    """
    for col, largura in larguras.items():
        ws.column_dimensions[col].width = largura


def criar_aba_clientes(wb: openpyxl.Workbook) -> None:
    """Cria ou recria a aba CLIENTES com dados de exemplo.

    Args:
        wb: Workbook aberto.
    """
    if "CLIENTES" in wb.sheetnames:
        del wb["CLIENTES"]
    ws = wb.create_sheet("CLIENTES", 0)
    ws.row_dimensions[1].height = 24

    cabecalho = ["CODIGO_CLIENTE", "NOME", "CONTATO", "OBSERVACOES"]
    ws.append(cabecalho)
    estilo_cabecalho(ws, 1, "0D47A1")

    exemplos = [
        ["DUDE", "Dude Brand", "contato@dude.com", "Cliente exemplo"],
        ["MR",   "Media Rats", "hello@mediarats.com", "Cliente interno"],
        ["TEST", "Teste Co",   "", "Dados de teste"],
    ]
    for row in exemplos:
        ws.append(row)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(vertical="center")
            cell.font = Font(name="Segoe UI", size=10)

    ajustar_colunas(ws, {"A": 18, "B": 28, "C": 28, "D": 32})
    ws.freeze_panes = "A2"


def criar_aba_conteudos(wb: openpyxl.Workbook) -> None:
    """Cria ou recria a aba CONTEUDOS com dados de exemplo.

    Args:
        wb: Workbook aberto.
    """
    if "CONTEUDOS" in wb.sheetnames:
        del wb["CONTEUDOS"]
    ws = wb.create_sheet("CONTEUDOS", 1)
    ws.row_dimensions[1].height = 30

    cabecalho = (
        ["PROTOCOLO", "CODIGO_CLIENTE", "CLIENTE", "NUMERO_SOLICITACAO",
         "TEMA", "STATUS", "DATA_PLANEJADA"]
        + [f"PROMPT {i}" for i in range(1, 11)]
    )
    ws.append(cabecalho)
    estilo_cabecalho(ws, 1, "1565C0")

    exemplos = [
        {
            "protocolo": "DUDE#1",
            "codigo": "DUDE",
            "cliente": "Dude Brand",
            "numero": 1,
            "tema": "Lançamento de Verão",
            "status": "Planejado",
            "data": "30/04/2025",
            "prompts": [
                "Cena de praia tropical ao entardecer com produto de bebida energética em destaque, estilo fotografia comercial vibrante",
                "Jovens surfistas celebrando vitória na praia, logotipo Dude Brand visível, paleta de cores azul e laranja",
                "Produto Dude Energy colocado em areia molhada com ondas ao fundo, efeito bokeh suave, cores quentes",
                "Flat lay de produtos de verão: óculos, protetor solar, produto Dude Brand, fundo de madeira rústica",
                "Homem atlético correndo na praia ao amanhecer, silhueta contra céu colorido, marca discreta no canto",
                "Coquetéis tropicais com lata Dude Brand, fundo de folhagens exóticas, estilo editorial de moda",
                "Vista aérea de piscina luxuosa com flutuadores coloridos e produto Dude visível, estilo lifestyle",
                "Casal jovem tomando Dude Energy ao pôr do sol na varanda com vista para o oceano",
                "Detalhe macro da lata Dude Brand com gotas de água, fundo desfocado de praia, alta qualidade",
                "Festival de música ao ar livre com bandeiras Dude Brand, público animado, iluminação colorida noturna",
            ],
        },
        {
            "protocolo": "MR#1",
            "codigo": "MR",
            "cliente": "Media Rats",
            "numero": 1,
            "tema": "Identidade Visual",
            "status": "Pendente",
            "data": "05/05/2025",
            "prompts": [
                "Logo Media Rats em estilo cyberpunk com rato geométrico neon em fundo preto, design moderno",
                "Escritório criativo futurista com telas holográficas mostrando dados de mídia social, equipe diversa",
                "Mockup de smartphone com interface de dashboard de analytics colorida, fundo gradiente roxo",
                "Ícone de rato vetorial minimalista em gradiente azul e roxo para uso em redes sociais",
                "Banner para LinkedIn com texto 'Media Rats - Criatividade Que Converte', fundo escuro profissional",
                "Colagem artística de elementos de marketing digital: gráficos, trends, emojis, fundo branco",
                "Mascote rato animado segurando tablet com gráficos de crescimento, estilo cartoon moderno",
                "Story do Instagram com paleta de cores da marca, tipografia bold, call-to-action vibrante",
                "Apresentação de pitch com slides escuros, dados de resultado de campanha, visual executivo",
                "Foto estilo editorial de mesa de trabalho criativa com notebook, câmera, caderno e café",
            ],
        },
    ]

    for ex in exemplos:
        linha = [
            ex["protocolo"], ex["codigo"], ex["cliente"], ex["numero"],
            ex["tema"], ex["status"], ex["data"],
        ] + ex["prompts"]
        ws.append(linha)
        for cell in ws[ws.max_row]:
            cell.alignment = Alignment(vertical="center", wrap_text=False)
            cell.font = Font(name="Segoe UI", size=10)

    ajustar_colunas(ws, {
        "A": 12, "B": 16, "C": 20, "D": 10,
        "E": 22, "F": 14, "G": 16,
        **{get_column_letter(i + 8): 55 for i in range(10)},
    })
    ws.freeze_panes = "H2"


def criar_aba_avaliacao(wb: openpyxl.Workbook) -> None:
    """Cria ou recria a aba AVALIACAO_DETALHADA vazia.

    Args:
        wb: Workbook aberto.
    """
    if "AVALIACAO_DETALHADA" in wb.sheetnames:
        del wb["AVALIACAO_DETALHADA"]
    ws = wb.create_sheet("AVALIACAO_DETALHADA", 2)
    ws.row_dimensions[1].height = 28

    cabecalho = [
        "Protocolo", "Cliente", "Tema", "Numero_Arte",
        "Prompt_Utilizado", "Imagem", "Qualidade",
        "Comentarios", "Usar?", "Data_Revisao", "Proximas_Acoes",
    ]
    ws.append(cabecalho)
    estilo_cabecalho(ws, 1, "2E7D32")

    ajustar_colunas(ws, {
        "A": 14, "B": 20, "C": 22, "D": 12,
        "E": 55, "F": 35, "G": 12,
        "H": 30, "I": 10, "J": 16, "K": 30,
    })
    ws.freeze_panes = "A2"


def main() -> None:
    """Cria ou atualiza a planilha com estrutura completa e dados de exemplo."""
    PLANILHA_PATH.parent.mkdir(parents=True, exist_ok=True)

    if PLANILHA_PATH.exists():
        print(f"Atualizando planilha existente: {PLANILHA_PATH}")
        wb = openpyxl.load_workbook(PLANILHA_PATH)
    else:
        print(f"Criando nova planilha: {PLANILHA_PATH}")
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    criar_aba_clientes(wb)
    criar_aba_conteudos(wb)
    criar_aba_avaliacao(wb)

    wb.save(PLANILHA_PATH)
    print(f"\n✅ Planilha salva com sucesso!\nAbas criadas: {wb.sheetnames}")
    print(f"\nCaminho: {PLANILHA_PATH}")


if __name__ == "__main__":
    main()
