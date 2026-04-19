"""
Testes de integração para excel/writer.py.
Usa planilhas temporárias em disco via fixture tmp_path do pytest.
"""

import pytest
import openpyxl
from pathlib import Path

from excel.reader import Solicitacao
from excel.writer import ExcelWriter


@pytest.fixture
def planilha_temp(tmp_path: Path) -> Path:
    """Cria uma planilha mínima de teste com cabeçalho e uma linha de dados."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CONTEUDOS"
    ws.append(["PROTOCOLO", "CLIENTE", "TEMA", "STATUS"])
    ws.append(["DUDE#1", "Dude Inc.", "Verão 2025", "Planejado"])
    caminho = tmp_path / "planilha_teste.xlsx"
    wb.save(caminho)
    return caminho


@pytest.fixture
def solicitacao_teste() -> Solicitacao:
    return Solicitacao(
        linha_excel=2,
        codigo_cliente="DUDE",
        cliente="Dude Inc.",
        numero_solicitacao=1,
        protocolo="DUDE#1",
        tema="Verão 2025",
        prompts=["Arte 1", "Arte 2"],
        status="Planejado",
    )


class TestAtualizarStatus:
    def test_status_escrito_na_planilha(self, planilha_temp, solicitacao_teste):
        writer = ExcelWriter(planilha_temp)
        writer.atualizar_status(solicitacao_teste, "Gerado", adicionar_data=False)

        wb = openpyxl.load_workbook(planilha_temp)
        ws = wb.active
        assert ws.cell(row=2, column=4).value == "Gerado"

    def test_cor_aplicada_na_celula(self, planilha_temp, solicitacao_teste):
        writer = ExcelWriter(planilha_temp)
        writer.atualizar_status(solicitacao_teste, "Gerado", adicionar_data=False)

        from utils.status import STATUS_CORES_EXCEL
        cor_esperada = STATUS_CORES_EXCEL["Gerado"]

        wb = openpyxl.load_workbook(planilha_temp)
        ws = wb.active
        fill = ws.cell(row=2, column=4).fill
        assert fill.fgColor.rgb.endswith(cor_esperada)

    def test_data_geracao_criada_quando_solicitada(self, planilha_temp, solicitacao_teste):
        writer = ExcelWriter(planilha_temp)
        writer.atualizar_status(solicitacao_teste, "Gerado", adicionar_data=True)

        wb = openpyxl.load_workbook(planilha_temp)
        ws = wb.active
        col_data = None
        for cell in ws[1]:
            if cell.value and str(cell.value).upper() == "DATA_GERACAO":
                col_data = cell.column
                break
        assert col_data is not None
        assert ws.cell(row=2, column=col_data).value is not None

    def test_status_invalido_usa_cor_branco(self, planilha_temp, solicitacao_teste):
        writer = ExcelWriter(planilha_temp)
        writer.atualizar_status(solicitacao_teste, "StatusInexistente", adicionar_data=False)

        wb = openpyxl.load_workbook(planilha_temp)
        ws = wb.active
        fill = ws.cell(row=2, column=4).fill
        assert fill.fgColor.rgb.endswith("FFFFFF")


class TestRegistrarConclusao:
    def test_status_atualizado_para_gerado(self, planilha_temp, solicitacao_teste):
        writer = ExcelWriter(planilha_temp)
        writer.registrar_conclusao(solicitacao_teste, ["output/DUDE#1_001.jpg"])

        wb = openpyxl.load_workbook(planilha_temp)
        ws = wb[wb.sheetnames[0]]
        assert ws.cell(row=2, column=4).value == "Gerado"

    def test_aba_avaliacao_criada(self, planilha_temp, solicitacao_teste):
        writer = ExcelWriter(planilha_temp)
        writer.registrar_conclusao(solicitacao_teste, ["output/DUDE#1_001.jpg"])

        wb = openpyxl.load_workbook(planilha_temp)
        nomes = [s.upper() for s in wb.sheetnames]
        assert "AVALIACAO_DETALHADA" in nomes
